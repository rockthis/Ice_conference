from openpyxl import load_workbook
from selenium import webdriver
from time import sleep
import xlrd
from selenium.common.exceptions import NoSuchElementException

def check_exists_by_xpath(xpath):
    try:
        driver.find_element_by_xpath(xpath)
    except NoSuchElementException:
        return False
    return True

workbook = xlrd.open_workbook('Ice_members.xlsx')
worksheet = workbook.sheet_by_index(0)
rows = worksheet.nrows
links = []
for i in range (0,rows - 1):
    links.append(worksheet.cell_value(1 + i,1))

print(len(links))

driver = webdriver.Chrome('C:\\Users\\Admin\\Downloads\\chromedriver')    #windows

mail = []
site = []
for el in links:
    driver.get(el)
    if check_exists_by_xpath('//*[@id="page-wrapper"]/div/div[1]/div/div[1]/div/div/div/div/div/div/div[1]/div/div[2]/div/a'):
        site.append(driver.find_element_by_xpath('//*[@id="page-wrapper"]/div/div[1]/div/div[1]/div/div/div/div/div/div/div[1]/div/div[2]/div/a').get_attribute('href'))
    else:
        site.append('No Site')
    if check_exists_by_xpath('//*[@id="page-wrapper"]/div/div[1]/div/div[1]/div/div/div/div/div/div/div[2]/div/div[2]/div/a'):
        mail.append(driver.find_element_by_xpath('//*[@id="page-wrapper"]/div/div[1]/div/div[1]/div/div/div/div/div/div/div[2]/div/div[2]/div/a').text)
        print(driver.find_element_by_xpath('//*[@id="page-wrapper"]/div/div[1]/div/div[1]/div/div/div/div/div/div/div[2]/div/div[2]/div/a').text)
    else:
        mail.append('No mail')
driver.quit()

wb = load_workbook(filename='Ice_members.xlsx')
sheet = wb.active
sheet.cell(row=1, column=6).value = 'Mail'
for i in range(0,len(site)):
    sheet.cell(row=2 + i, column=3).value = site[i]
for i in range(0,len(mail)):
    sheet.cell(row=2 + i, column=6).value = mail[i]
wb.save('Ice_members.xlsx')