from openpyxl import Workbook
from selenium import webdriver
from time import sleep
from selenium.common.exceptions import NoSuchElementException

# driver = webdriver.Chrome('/Users/kostyafrolov/Downloads/chromedriver')   mac

driver = webdriver.Chrome('C:\\Users\\Admin\\Downloads\\chromedriver')    #windows

driver.get('http://www.icetotallygaming.com/exhibitor-list')


sleep(3)


def check_exists_by_xpath(xpath):
    try:
        driver.find_element_by_xpath(xpath)
    except NoSuchElementException:
        return False
    return True

name = []
link_ICE = []

while True:
    descriptions = driver.find_elements_by_xpath('//*[@class="exhibitor-description"]/p/a')
    for el in descriptions:
        name.append(el.find_element_by_xpath('../../../div[@class="exhibitor-title"]').text)
        link_ICE.append(el.get_attribute('href'))
    if check_exists_by_xpath('//*[@title="Go to next page"]'):
        driver.find_element_by_xpath('//*[@title="Go to next page"]').click()
        sleep(2)
    else:
        break

wb = Workbook()
sheet = wb.active
sheet.cell(row=1, column=1).value = 'Name'
sheet.cell(row=1, column=2).value = 'Link in Ice'
sheet.cell(row=1, column=3).value = 'Url'
sheet.cell(row=1, column=4).value = 'Similar Web Rating'
sheet.cell(row=1, column=5).value = 'Activity'
for i in range(0,len(name)):
    sheet.cell(row=2 + i, column=1).value = name[i]
for i in range(0,len(link_ICE)):
    sheet.cell(row=2 + i, column=2).value = link_ICE[i]

wb.save('Ice_members.xlsx')


driver.quit()